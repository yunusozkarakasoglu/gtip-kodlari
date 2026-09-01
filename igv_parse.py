#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""İGV (İlave Gümrük Vergisi) listelerini parse et — Ek-1, Ek-2, Ek-3.
Kaynak: ithalat.ticaret.gov.tr igv_2026.zip (resmi)
Çıktı: Masaüstü/gtip_kodlari/igv_2026.json
"""
import openpyxl, re, json, os

OUT = "/home/yunus/Masaüstü/gtip_kodlari"
KAYNAK = f"{OUT}/igv_2026"

def sayi(v):
    if v in (None, ""):
        return None
    m = re.match(r"^(\d+(?:[.,]\d+)?)", str(v).strip())
    return float(m.group(1).replace(",", ".")) if m else None

def temiz_kod(v):
    if not v:
        return None
    s = str(v).strip().replace(".", "").replace(" ", "")
    return s if s.isdigit() else None

def parse_ek1():
    wb = openpyxl.load_workbook(f"{KAYNAK}/Ek-1.xlsx")
    ws = wb["EK-1"]
    sonuc = []
    for r in range(6, ws.max_row + 1):
        g = temiz_kod(ws.cell(r, 1).value)
        if not g:
            continue
        dip = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ""
        sonuc.append({
            "liste": "EK-1", "gtip": g, "dipnot": dip,
            "ab": sayi(ws.cell(r, 3).value),      # sütun 1: AB+EFTA+STA
            "kol2": sayi(ws.cell(r, 4).value),    # sütun 2
            "kol3": sayi(ws.cell(r, 5).value),    # sütun 3
            "gts": sayi(ws.cell(r, 6).value),     # GTS ülkeleri (4/5/6 eşit)
            "diger": sayi(ws.cell(r, 9).value),   # sütun 7 = D.Ü.
        })
    return sonuc

def parse_ek2():
    wb = openpyxl.load_workbook(f"{KAYNAK}/Ek-2 Ek-3.xlsx")
    ws = wb["Nihai İGV Ek 2"]
    sonuc = []
    for r in range(6, ws.max_row + 1):
        g = temiz_kod(ws.cell(r, 1).value)
        if not g:
            continue
        dip = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ""
        sonuc.append({
            "liste": "EK-2", "gtip": g, "dipnot": dip,
            "ab": sayi(ws.cell(r, 3).value),       # AB, B-HER, BK, EFTA, F.ADA, G.KORE, M...
            "kos": sayi(ws.cell(r, 4).value),      # Kosova
            "sng": sayi(ws.cell(r, 5).value),      # Singapur
            "vnz": sayi(ws.cell(r, 6).value),      # Venezuela
            "bae": sayi(ws.cell(r, 7).value),      # BAE
            "gts_eagu": sayi(ws.cell(r, 8).value), # GTS: En Az Gelişmiş
            "gts_otdu": sayi(ws.cell(r, 9).value), # GTS: Özel Teşvik
            "gts_gyu": sayi(ws.cell(r, 10).value), # GTS: Gelişme Yolunda
            "diger": sayi(ws.cell(r, 11).value),   # D.Ü.
        })
    return sonuc

def parse_ek3():
    wb = openpyxl.load_workbook(f"{KAYNAK}/Ek-2 Ek-3.xlsx")
    ws = wb["Nihai İGV Ek 3"]
    sonuc = []
    for r in range(6, ws.max_row + 1):
        g = temiz_kod(ws.cell(r, 1).value)
        if not g:
            continue
        dip = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ""
        sonuc.append({
            "liste": "EK-3", "gtip": g, "dipnot": dip,
            "ab": sayi(ws.cell(r, 3).value),        # AB, BK, B-HER, EFTA, F.ADA
            "g_kore": sayi(ws.cell(r, 4).value),
            "mlz": sayi(ws.cell(r, 5).value),
            "sng": sayi(ws.cell(r, 6).value),
            "kos": sayi(ws.cell(r, 7).value),
            "iran": sayi(ws.cell(r, 8).value),
            "vnz": sayi(ws.cell(r, 9).value),
            "bae": sayi(ws.cell(r, 10).value),
            "gts_eagu": sayi(ws.cell(r, 11).value),
            "gts_otdu": sayi(ws.cell(r, 12).value),
            "gts_gyu": sayi(ws.cell(r, 13).value),
            "diger": sayi(ws.cell(r, 14).value),    # D.Ü.
        })
    return sonuc

def main():
    e1 = parse_ek1()
    e2 = parse_ek2()
    e3 = parse_ek3()
    tum = e1 + e2 + e3
    with open(f"{OUT}/igv_2026.json", "w", encoding="utf-8") as f:
        json.dump(tum, f, ensure_ascii=False, indent=1)
    print(f"✅ İGV: EK-1={len(e1)}, EK-2={len(e2)}, EK-3={len(e3)} → toplam {len(tum)}")
    # örnekler
    for x in tum:
        if x["gtip"].startswith("8408") or x["gtip"].startswith("8409"):
            print("  örnek:", x)
            break
    # 12 haneli mi kontrol
    uzun = {}
    for x in tum:
        uzun[len(x["gtip"])] = uzun.get(len(x["gtip"]), 0) + 1
    print("Kod uzunlukları:", dict(sorted(uzun.items())))

if __name__ == "__main__":
    main()
